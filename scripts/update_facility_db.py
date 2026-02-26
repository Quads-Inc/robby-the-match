#!/usr/bin/env python3
"""
施設基準届出受理名簿（令和8年2月作成・令和8年1月1日現在）を元に
施設データベースを更新するスクリプト

データソース:
- 関東信越厚生局 施設基準届出受理名簿 項目別1（入院基本料等）
- 関東信越厚生局 施設基準届出受理名簿 項目別2（特定入院料）
- 医療情報ネット CSV（2025年12月1日時点）

出力:
- kanagawa_ward_data.json の更新
- worker_facilities.js の features/beds 更新用データ
"""

import openpyxl
import json
import csv
import os
import re
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'data', 'public_data')

# Target cities for 神奈川県西部
TARGET_CITIES = [
    '小田原', '秦野', '平塚', '藤沢', '茅ヶ崎', '茅ケ崎', '南足柄', '伊勢原',
    '大磯', '二宮', '中井', '大井', '松田', '山北', '開成', '箱根',
    '真鶴', '湯河原', '厚木'
]

# Nursing ratio mapping
NURSING_RATIO_MAP = {
    '急性期一般入院料１': '7:1',
    '急性期一般入院料２': '10:1',
    '急性期一般入院料３': '10:1',
    '急性期一般入院料４': '10:1',
    '急性期一般入院料５': '10:1',
    '急性期一般入院料６': '10:1',
    '地域一般入院料１': '13:1',
    '地域一般入院料２': '15:1',
    '地域一般入院料３': '15:1',
    '特別入院基本料': '—',
    '７対１入院基本料': '7:1',
    '１０対１入院基本料': '10:1',
    '１３対１入院基本料': '13:1',
    '１５対１入院基本料': '15:1',
}

# Separate mapping for entries that need context (療養 vs 回復期)
CONTEXTUAL_RATIO_MAP = {
    '療養病棟入院基本料': {'入院料１': '20:1', '入院料２': '25:1'},
    '回復期リハビリテーション病棟入院料１': '13:1',
    '回復期リハビリテーション病棟入院料２': '13:1',
    '回復期リハビリテーション病棟入院料３': '15:1',
    '回復期リハビリテーション病棟入院料４': '15:1',
    '回復期リハビリテーション病棟入院料５': '15:1',
    '地域包括ケア病棟入院料': '13:1',
    '地域包括医療病棟入院料': '10:1',
    '精神科急性期治療病棟入院料１': '13:1',
    '精神科急性期治療病棟入院料２': '15:1',
    '精神療養病棟入院料': '15:1',
    '認知症治療病棟入院料１': '20:1',
    '認知症治療病棟入院料２': '30:1',
    '緩和ケア病棟入院料１': '7:1',
    '緩和ケア病棟入院料２': '10:1',
    '障害者施設等入院基本料': None,  # ratio is in kubun
    '救命救急入院料': None,
    'ハイケアユニット入院医療管理料': None,
    '特殊疾患病棟入院料': '15:1',
    '小児入院医療管理料': None,
}

def parse_shisetsu_kijun():
    """Parse 施設基準 Excel files and extract hospital data"""

    files = [
        (os.path.join(DATA_DIR, 'shisetsu_kijun', 'koumoku1', 'koumokubetsu1_r0802',
         '14届出受理医療機関名簿（項目別1：医）神奈川r0802.xlsx'), '入院基本料等'),
        (os.path.join(DATA_DIR, 'shisetsu_kijun', 'koumoku2', 'koumokubetsu2_r0802',
         '14届出受理医療機関名簿（項目別2：医）神奈川r0802.xlsx'), '特定入院料'),
    ]

    hospitals = {}

    for filepath, category in files:
        if not os.path.exists(filepath):
            print(f"WARNING: {filepath} not found, skipping")
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[7]:
                continue
            name = str(row[7]).strip()
            address = str(row[9]) if row[9] else ''
            todoke = str(row[3]) if row[3] else ''
            med_code = str(row[4]) if row[4] else ''
            beds_str = str(row[12]) if row[12] else ''
            biko_key = str(row[17]).strip() if row[17] else ''
            biko_val = str(row[18]).strip() if row[18] else ''

            # Check if in target area
            in_target = any(city in address for city in TARGET_CITIES)
            if not in_target:
                continue

            if name not in hospitals:
                hospitals[name] = {
                    'name': name,
                    'address': address,
                    'med_code': med_code,
                    'beds_raw': beds_str,
                    'entries': []
                }

            # Parse structured rows
            if not biko_key and todoke:
                # Main entry row - start new entry
                hospitals[name]['entries'].append({
                    'category': category,
                    'todoke': todoke,
                    'details': {}
                })
            elif biko_key and hospitals[name]['entries']:
                key = biko_key.rstrip(':：')
                hospitals[name]['entries'][-1]['details'][key] = biko_val

        wb.close()

    return hospitals


def extract_ward_info(hospitals):
    """Extract structured ward information from parsed data"""

    result = {}

    for name, data in hospitals.items():
        wards = []
        admission_fees = []
        primary_ratio = None

        for entry in data['entries']:
            todoke = entry['todoke']
            details = entry['details']

            # Skip non-ward entries (有床診療所, etc.)
            if '有床診療所' in todoke:
                continue

            # Only process entries with 入院 or 病棟
            if '入院' not in todoke and '病棟' not in todoke and 'ケア' not in todoke:
                continue

            kubun = details.get('区分', '')
            beds_str = details.get('病床数', '')
            ward_type = details.get('病棟種別', '')
            ward_count = details.get('病棟数', '')

            # Parse bed count
            beds = 0
            if beds_str:
                m = re.search(r'(\d+)', beds_str)
                if m:
                    beds = int(m.group(1))

            # Determine nursing ratio
            nursing_ratio = None
            # First check kubun (区分) for basic ratios
            if kubun:
                for key, ratio in NURSING_RATIO_MAP.items():
                    if key in kubun:
                        nursing_ratio = ratio
                        break
            # Then check todoke name for contextual ratios
            if not nursing_ratio:
                for key, ratio in CONTEXTUAL_RATIO_MAP.items():
                    if key in todoke:
                        if isinstance(ratio, dict):
                            # Need kubun to determine ratio (e.g., 療養病棟)
                            for sub_key, sub_ratio in ratio.items():
                                if sub_key in kubun:
                                    nursing_ratio = sub_ratio
                                    break
                        elif ratio:
                            nursing_ratio = ratio
                        break
            # Fallback: check kubun in basic map
            if not nursing_ratio and kubun:
                for key, ratio in NURSING_RATIO_MAP.items():
                    if key in todoke:
                        nursing_ratio = ratio
                        break

            # Determine ward function
            function = '急性期'
            if '回復期' in todoke:
                function = '回復期'
            elif '療養' in todoke:
                function = '慢性期'
            elif '地域包括' in todoke:
                function = '急性期'  # 地域包括ケア is acute-phase classification
            elif '精神' in todoke or '認知症' in todoke:
                function = '精神科'
            elif '緩和' in todoke:
                function = '慢性期'
            elif '障害者' in todoke:
                function = '慢性期'
            elif '救命救急' in todoke or 'ハイケアユニット' in todoke:
                function = '高度急性期'
            elif '地域一般' in kubun:
                function = '急性期'

            # Build admission fee name
            fee_name = todoke
            if kubun:
                fee_name = f"{todoke}（{kubun}）"

            if fee_name not in admission_fees:
                admission_fees.append(fee_name)

            # Determine primary nursing ratio (from 一般病棟入院基本料)
            if '一般病棟入院基本料' in todoke and nursing_ratio and not primary_ratio:
                primary_ratio = nursing_ratio

            ward = {
                'todoke': todoke,
                'kubun': kubun,
                'function': function,
                'beds': beds,
                'nursing_ratio': nursing_ratio,
                'ward_type': ward_type,
                'fee_name': fee_name,
            }

            # Only add if it has useful info
            if beds > 0 or kubun:
                wards.append(ward)

        if not wards:
            continue

        # Parse total beds from beds_raw
        total_general = 0
        total_therapy = 0
        total_mental = 0
        if data['beds_raw']:
            m_gen = re.search(r'一般\s*(\d+)', data['beds_raw'])
            m_ther = re.search(r'療養\s*(\d+)', data['beds_raw'])
            m_ment = re.search(r'精神\s*(\d+)', data['beds_raw'])
            if m_gen: total_general = int(m_gen.group(1))
            if m_ther: total_therapy = int(m_ther.group(1))
            if m_ment: total_mental = int(m_ment.group(1))

        total_beds = total_general + total_therapy + total_mental

        # Build nursing ratio string
        if primary_ratio:
            ratio_str = primary_ratio
        elif wards:
            # Use first non-null ratio
            for w in wards:
                if w['nursing_ratio']:
                    ratio_str = w['nursing_ratio']
                    break
            else:
                ratio_str = '不明'
        else:
            ratio_str = '不明'

        result[name] = {
            'name': name,
            'address': data['address'],
            'med_code': data['med_code'],
            'total_beds': total_beds,
            'general_beds': total_general,
            'therapy_beds': total_therapy,
            'mental_beds': total_mental,
            'nursing_ratio': ratio_str,
            'admission_fees': admission_fees,
            'wards': wards,
            'data_source': '施設基準届出受理名簿R8.1.1+医療情報ネット2025.12',
            'data_date': '2026-01-01',
        }

    return result


def update_ward_data(new_data):
    """Update kanagawa_ward_data.json with new data"""

    ward_file = os.path.join(DATA_DIR, 'kanagawa_ward_data.json')

    # Read existing data
    with open(ward_file, 'r', encoding='utf-8') as f:
        existing = json.load(f)

    updated_count = 0
    added_count = 0

    for full_name, data in new_data.items():
        # Find matching hospital in existing data (by partial name match)
        short_name = None
        for existing_name in existing:
            if existing_name in full_name or full_name in existing_name:
                short_name = existing_name
                break

        # Build ward list in existing format
        new_wards = []
        new_fees = []
        functions = set()

        for w in data['wards']:
            ratio_suffix = f"（{w['nursing_ratio']}）" if w['nursing_ratio'] else ''
            ward_name = w['todoke'].replace('入院基本料', '').replace('入院料', '')
            if w['kubun']:
                ward_name += f"（{w['kubun']}）"

            new_wards.append({
                'name': f"{ward_name}{ratio_suffix}" if ratio_suffix and ratio_suffix not in ward_name else ward_name,
                'function': w['function'],
                'admissionFee': w['fee_name'],
                'beds': w['beds'],
                'nursesFT': 0,  # Not available from 施設基準
                'nursesPT': 0,
            })

            if w['fee_name'] not in new_fees:
                new_fees.append(w['fee_name'])
            functions.add(w['function'])

        # Build nursing ratio string
        ratio_parts = []
        for w in data['wards']:
            if w['nursing_ratio'] and w['beds'] > 0:
                ward_short = ''
                if '一般病棟' in w['todoke']:
                    ward_short = '一般'
                elif '回復期' in w['todoke']:
                    ward_short = '回復期'
                elif '療養' in w['todoke']:
                    ward_short = '療養'
                elif '障害者' in w['todoke']:
                    ward_short = '障害者'
                elif '地域一般' in w.get('kubun', ''):
                    ward_short = '地域一般'
                elif '精神' in w['todoke']:
                    ward_short = '精神'
                elif '特定機能' in w['todoke']:
                    ward_short = '特定機能'

                if ward_short:
                    ratio_parts.append(f"{ward_short}{w['nursing_ratio']}")

        nursing_ratio = '/'.join(ratio_parts) if ratio_parts else data['nursing_ratio']

        entry = {
            'name': data['name'],
            'cityName': '',  # Will be filled from address
            'nursingRatio': nursing_ratio,
            'admissionFees': new_fees,
            'wardCount': len(new_wards),
            'functions': list(functions),
            'wards': new_wards,
            'dataSource': data['data_source'],
            'dataDate': data['data_date'],
        }

        # Extract city from address
        m = re.search(r'((?:小田原|平塚|秦野|藤沢|茅ヶ崎|茅ケ崎|厚木|南足柄|伊勢原)市|(?:大磯|二宮|中井|大井|松田|山北|開成|箱根|真鶴|湯河原)町)', data['address'])
        if m:
            entry['cityName'] = m.group(1)

        if short_name:
            # Preserve existing nurse counts if available
            old = existing[short_name]
            if old.get('totalWardNurses'):
                entry['totalWardNurses'] = old['totalWardNurses']
            if old.get('medicalCode'):
                entry['medicalCode'] = old['medicalCode']

            # Update ward nurse counts from existing data if ward names partially match
            for nw in entry['wards']:
                for ow in old.get('wards', []):
                    if ow['function'] == nw['function'] and ow.get('nursesFT', 0) > 0:
                        nw['nursesFT'] = ow['nursesFT']
                        nw['nursesPT'] = ow.get('nursesPT', 0)
                        break

            existing[short_name] = entry
            updated_count += 1
        else:
            # Add new hospital
            # Use a short name
            simple_name = full_name
            for prefix in ['医療法人社団', '医療法人財団', '医療法人', '特定医療法人社団',
                          '特定医療法人', '社会福祉法人', '独立行政法人', '一般財団法人',
                          '公益財団法人', '国家公務員共済組合連合会']:
                simple_name = simple_name.replace(prefix, '').strip()
            simple_name = re.sub(r'^[\s　]+', '', simple_name)

            existing[simple_name] = entry
            added_count += 1

    # Save updated file
    with open(ward_file, 'w', encoding='utf-8') as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"Updated: {updated_count} hospitals, Added: {added_count} new hospitals")
    print(f"Total hospitals in DB: {len(existing)}")
    return existing


def generate_features_update(new_data, ward_data):
    """Generate update data for worker_facilities.js"""

    updates = {}

    for name, data in new_data.items():
        # Build features string
        parts = []

        for w in data['wards']:
            if w['beds'] > 0 and w['nursing_ratio']:
                ward_type = ''
                if '一般病棟' in w['todoke']:
                    ward_type = '一般病棟'
                elif '回復期' in w['todoke']:
                    ward_type = '回復期'
                elif '療養' in w['todoke']:
                    ward_type = '療養病棟'
                elif '障害者' in w['todoke']:
                    ward_type = '障害者病棟'
                elif '地域一般' in w.get('kubun', ''):
                    ward_type = '地域一般'

                if ward_type:
                    kubun = w.get('kubun', '')
                    if kubun:
                        parts.append(f"{ward_type}{w['nursing_ratio']}({kubun})")
                    else:
                        parts.append(f"{ward_type}{w['nursing_ratio']}")

        if parts:
            updates[name] = {
                'nursing_info': '・'.join(parts),
                'total_beds': data['total_beds'],
                'wards_summary': parts,
            }

    return updates


def main():
    print("=" * 60)
    print("施設データベース更新スクリプト")
    print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Step 1: Parse 施設基準
    print("\n[1] 施設基準データ解析中...")
    raw_data = parse_shisetsu_kijun()
    print(f"   {len(raw_data)} hospitals found in target area")

    # Step 2: Extract ward info
    print("\n[2] 病棟情報抽出中...")
    ward_info = extract_ward_info(raw_data)
    print(f"   {len(ward_info)} hospitals with ward data")

    # Step 3: Update kanagawa_ward_data.json
    print("\n[3] kanagawa_ward_data.json 更新中...")
    updated_ward_data = update_ward_data(ward_info)

    # Step 4: Generate features update data
    print("\n[4] worker_facilities.js 更新データ生成中...")
    features_updates = generate_features_update(ward_info, updated_ward_data)

    # Save features update as reference
    features_file = os.path.join(DATA_DIR, 'features_update_r0802.json')
    with open(features_file, 'w', encoding='utf-8') as f:
        json.dump(features_updates, f, ensure_ascii=False, indent=2)
    print(f"   Features update data saved to {features_file}")
    print(f"   {len(features_updates)} hospitals with nursing info updates")

    # Print summary of key hospitals
    print("\n" + "=" * 60)
    print("主要病院サマリ")
    print("=" * 60)

    key_hospitals = ['小林病院', '小田原市立病院', '間中病院', '箱根リハビリテーション病院',
                     '秦野赤十字病院', '平塚共済病院', '茅ヶ崎市立病院', '藤沢市民病院',
                     '東海大学医学部付属病院', '伊勢原協同病院', '神奈川県立足柄上病院']

    for key in key_hospitals:
        for name, data in ward_info.items():
            if key in name:
                wards_str = ' / '.join([
                    f"{w['todoke']}{'(' + w['kubun'] + ')' if w['kubun'] else ''} {w['beds']}床 [{w['nursing_ratio'] or '—'}]"
                    for w in data['wards'] if w['beds'] > 0
                ])
                print(f"\n{data['name']}（{data['total_beds']}床）")
                print(f"  {wards_str}")
                break

    print("\n\n完了!")


if __name__ == '__main__':
    main()
