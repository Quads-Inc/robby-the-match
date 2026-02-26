#!/usr/bin/env python3
"""
parse_bed_function.py - 病床機能報告 令和5年度 Excel parser

Parses the official 病床機能報告 (Bed Function Report) Excel files
to extract ward-level data for 神奈川県西部 hospitals, and compares
with the existing kanagawa_ward_data.json.

Data source: 厚生労働省 病床機能報告 令和5年度
- bed_function_facility.xlsx: Facility-level data (施設票)
- bed_function_ward_kanto.xlsx: Ward-level data for Kanto region (病棟票)

Usage:
    python3 scripts/parse_bed_function.py
"""

import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# =============================================================================
# Configuration
# =============================================================================

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data", "public_data")

WARD_FILE = os.path.join(DATA_DIR, "bed_function_ward_kanto.xlsx")
FACILITY_FILE = os.path.join(DATA_DIR, "bed_function_facility.xlsx")
EXISTING_JSON = os.path.join(DATA_DIR, "kanagawa_ward_data.json")
OUTPUT_JSON = os.path.join(DATA_DIR, "bed_function_parsed.json")
COMPARISON_JSON = os.path.join(DATA_DIR, "bed_function_comparison.json")

KANAGAWA_PREF_CODE = "14"

# 神奈川県西部 target cities (市区町村名)
NISHI_KANAGAWA_CITIES = {
    "小田原市", "秦野市", "平塚市", "藤沢市", "茅ヶ崎市",
    "南足柄市", "伊勢原市", "大磯町", "二宮町", "中井町",
    "大井町", "松田町", "山北町", "開成町", "箱根町",
    "真鶴町", "湯河原町",
    # Also include nearby cities that appear in existing data
    "厚木市", "海老名市", "寒川町",
}

# Ward file column indices (0-based, from row 5 header)
# Row 5 is the clean merged header row; data starts at row 7
WARD_COLS = {
    "medical_code":     0,   # オープンデータ医療機関コード
    "hospital_name":    1,   # 医療機関名
    "pref_code":        2,   # 都道府県コード
    "secondary_area_code": 3,  # 二次医療圏コード
    "secondary_area_name": 4,  # 二次医療圏名
    "vision_area_code": 5,   # 構想区域コード
    "vision_area_name": 6,   # 構想区域名称
    "city_code":        7,   # 市区町村コード
    "city_name":        8,   # 市区町村名称
    "med_code_ika":     9,   # 医療機関コード（医科）
    "med_code_shika":   10,  # 医療機関コード（歯科）
    "ward_code":        11,  # 病棟コード
    "ward_name":        12,  # 病棟名
    "build_year":       13,  # 建築時期
    "structure":        14,  # 構造
    "function_r5":      15,  # 2023年7月1日時点の機能
    "function_r7":      16,  # 2025年7月1日の予定機能
    "transition_plan":  17,  # 移行予定先
    "general_beds":     18,  # 一般病床_許可病床
    "general_max":      19,  # 一般病床_最大使用
    "general_min":      20,  # 一般病床_最小使用
    "general_keika":    21,  # 一般病床_経過措置
    "therapy_beds":     22,  # 療養病床_許可病床
    "therapy_max":      23,  # 療養病床_最大使用
    "therapy_min":      24,  # 療養病床_最小使用
    "therapy_med_beds": 25,  # 療養_医療療養_許可
    "therapy_med_max":  26,  # 療養_医療療養_最大使用
    "therapy_med_min":  27,  # 療養_医療療養_最小使用
    "therapy_care_beds": 28, # 療養_介護療養_許可
    "therapy_care_max": 29,  # 療養_介護療養_最大使用
    "therapy_care_min": 30,  # 療養_介護療養_最小使用
    "zero_reason":      31,  # 最大使用病床数0の理由
    "admission_fee":    32,  # 算定する入院基本料・特定入院料
    "admission_fee_beds": 33,  # 届出病床数
    "special_fee_1":    34,  # 病室単位の特定入院料1
    "special_fee_1_beds": 35,  # 病室単位の特定入院料1_届出病床数
    "special_fee_2":    36,  # 病室単位の特定入院料2
    "special_fee_2_beds": 37,  # 病室単位の特定入院料2_届出病床数
    "care_therapy":     38,  # 介護療養施設サービス費の届出
    "nurse_ft":         39,  # 看護師_常勤
    "nurse_pt":         40,  # 看護師_非常勤
    "prac_nurse_ft":    41,  # 准看護師_常勤
    "prac_nurse_pt":    42,  # 准看護師_非常勤
    "aide_ft":          43,  # 看護補助者_常勤
    "aide_pt":          44,  # 看護補助者_非常勤
    "midwife_ft":       45,  # 助産師_常勤
    "midwife_pt":       46,  # 助産師_非常勤
    "pt_ft":            47,  # 理学療法士_常勤
    "pt_pt":            48,  # 理学療法士_非常勤
    "ot_ft":            49,  # 作業療法士_常勤
    "ot_pt":            50,  # 作業療法士_非常勤
    "st_ft":            51,  # 言語聴覚士_常勤
    "st_pt":            52,  # 言語聴覚士_非常勤
    "pharmacist_ft":    53,  # 薬剤師_常勤
    "pharmacist_pt":    54,  # 薬剤師_非常勤
    "ce_ft":            55,  # 臨床工学技士_常勤
    "ce_pt":            56,  # 臨床工学技士_非常勤
    "dietitian_ft":     57,  # 管理栄養士_常勤
    "dietitian_pt":     58,  # 管理栄養士_非常勤
    "emt_ft":           59,  # 救急救命士_常勤
    "emt_pt":           60,  # 救急救命士_非常勤
    "main_dept":        61,  # 主とする診療科
    "dept_1":           62,  # 上位診療科_1位
    "dept_2":           63,  # 上位診療科_2位
    "dept_3":           64,  # 上位診療科_3位
    "new_patients_yr":  65,  # 新規入棟患者数（年間）
}

# Facility file column indices (0-based, from row 5)
FACILITY_COLS = {
    "type":             0,   # 病診区分
    "medical_code":     1,   # オープンデータ医療機関コード
    "hospital_name":    2,   # 医療機関名
    "pref_code":        3,   # 都道府県コード
    "secondary_area_code": 4,
    "secondary_area_name": 5,
    "vision_area_code": 6,
    "vision_area_name": 7,
    "city_code":        8,
    "city_name":        9,
    "med_code_ika":     10,
    "med_code_shika":   11,
    "owner":            12,  # 設置主体
    "dpc_group":        13,  # DPC群の種類
    "tokutei":          14,  # 特定機能病院
    "chiiki_shien":     15,  # 地域医療支援病院
    "sogo_nyuin":       16,  # 総合入院体制加算
    "kyuseiki_juujitsu": 17, # 急性期充実体制加算
    "seishin_juujitsu": 18,  # 精神科充実体制加算
    "zaitaku_shien":    19,  # 在宅療養支援病院
    "zaitaku_kouhou":   20,  # 在宅療養後方支援病院
}


def safe_int(val, default=0):
    """Safely convert value to int."""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def safe_float(val, default=0.0):
    """Safely convert value to float."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    """Safely convert value to string."""
    if val is None:
        return default
    return str(val).strip()


def estimate_nursing_ratio(admission_fee, general_beds, therapy_beds, nurse_ft, nurse_pt):
    """
    Estimate the nursing ratio from admission fee name and staffing.
    Returns a string like "7:1", "10:1", "13:1", "15:1", "20:1", or "不明".
    """
    fee = safe_str(admission_fee)
    if not fee or fee == "-":
        return "不明"

    # Direct ratio extraction from admission fee name
    ratio_map = {
        "急性期一般入院料１": "7:1",
        "急性期一般入院料２": "10:1",
        "急性期一般入院料３": "10:1",
        "急性期一般入院料４": "10:1",
        "急性期一般入院料５": "10:1",
        "急性期一般入院料６": "10:1",
        "地域一般入院料１": "13:1",
        "地域一般入院料２": "15:1",
        "地域一般入院料３": "15:1",
        "地域包括ケア病棟入院料１": "13:1",
        "地域包括ケア病棟入院料２": "13:1",
        "地域包括ケア病棟入院料３": "13:1",
        "地域包括ケア病棟入院料４": "13:1",
    }

    for key, ratio in ratio_map.items():
        if key in fee:
            return ratio

    # Recovery ward
    if "回復期" in fee:
        return "13:1" if "１" in fee or "２" in fee else "15:1"

    # Therapy ward
    if "療養病棟入院料" in fee:
        return "20:1"

    # Calculate from staffing if possible
    total_beds = general_beds + therapy_beds
    total_nurses = nurse_ft + nurse_pt * 0.5
    if total_beds > 0 and total_nurses > 0:
        ratio = total_beds / total_nurses
        if ratio <= 8:
            return "7:1"
        elif ratio <= 11:
            return "10:1"
        elif ratio <= 14:
            return "13:1"
        elif ratio <= 16:
            return "15:1"
        elif ratio <= 25:
            return "20:1"
        else:
            return f"推定{int(round(ratio))}:1"

    return "不明"


def parse_ward_file():
    """Parse the ward-level Excel file for 神奈川県西部 hospitals."""
    print(f"Loading ward file: {WARD_FILE}")
    print("  (This may take a moment for a 10MB file...)")

    wb = openpyxl.load_workbook(WARD_FILE, read_only=True)
    ws = wb.active

    hospitals = defaultdict(lambda: {
        "medical_code": "",
        "name": "",
        "city_code": "",
        "city_name": "",
        "secondary_area_name": "",
        "vision_area_name": "",
        "wards": [],
    })

    row_count = 0
    kanagawa_count = 0
    target_count = 0

    # Data starts at row 7 (1-indexed), so min_row=7
    for row in ws.iter_rows(min_row=7, values_only=True):
        row_count += 1
        pref_code = safe_str(row[WARD_COLS["pref_code"]])

        if pref_code != KANAGAWA_PREF_CODE:
            continue

        kanagawa_count += 1
        city_name = safe_str(row[WARD_COLS["city_name"]])

        if city_name not in NISHI_KANAGAWA_CITIES:
            continue

        target_count += 1

        medical_code = safe_str(row[WARD_COLS["medical_code"]])
        hospital_name = safe_str(row[WARD_COLS["hospital_name"]])

        # Update hospital info
        h = hospitals[medical_code]
        h["medical_code"] = medical_code
        h["name"] = hospital_name
        h["city_code"] = safe_str(row[WARD_COLS["city_code"]])
        h["city_name"] = city_name
        h["secondary_area_name"] = safe_str(row[WARD_COLS["secondary_area_name"]])
        h["vision_area_name"] = safe_str(row[WARD_COLS["vision_area_name"]])

        # Extract ward data
        ward_name = safe_str(row[WARD_COLS["ward_name"]])
        function_r5 = safe_str(row[WARD_COLS["function_r5"]])
        function_r7 = safe_str(row[WARD_COLS["function_r7"]])
        admission_fee = safe_str(row[WARD_COLS["admission_fee"]])
        admission_fee_beds = safe_int(row[WARD_COLS["admission_fee_beds"]])

        general_beds = safe_int(row[WARD_COLS["general_beds"]])
        general_max = safe_int(row[WARD_COLS["general_max"]])
        general_min = safe_int(row[WARD_COLS["general_min"]])
        therapy_beds = safe_int(row[WARD_COLS["therapy_beds"]])
        therapy_max = safe_int(row[WARD_COLS["therapy_max"]])
        therapy_min = safe_int(row[WARD_COLS["therapy_min"]])
        therapy_med_beds = safe_int(row[WARD_COLS["therapy_med_beds"]])

        nurse_ft = safe_float(row[WARD_COLS["nurse_ft"]])
        nurse_pt = safe_float(row[WARD_COLS["nurse_pt"]])
        prac_nurse_ft = safe_float(row[WARD_COLS["prac_nurse_ft"]])
        prac_nurse_pt = safe_float(row[WARD_COLS["prac_nurse_pt"]])
        aide_ft = safe_float(row[WARD_COLS["aide_ft"]])
        aide_pt = safe_float(row[WARD_COLS["aide_pt"]])

        # Rehab staff
        pt_ft = safe_float(row[WARD_COLS["pt_ft"]])
        pt_pt = safe_float(row[WARD_COLS["pt_pt"]])
        ot_ft = safe_float(row[WARD_COLS["ot_ft"]])
        ot_pt = safe_float(row[WARD_COLS["ot_pt"]])
        st_ft = safe_float(row[WARD_COLS["st_ft"]])
        st_pt = safe_float(row[WARD_COLS["st_pt"]])

        # Special fees
        special_fee_1 = safe_str(row[WARD_COLS["special_fee_1"]])
        special_fee_1_beds = safe_int(row[WARD_COLS["special_fee_1_beds"]])
        special_fee_2 = safe_str(row[WARD_COLS["special_fee_2"]])
        special_fee_2_beds = safe_int(row[WARD_COLS["special_fee_2_beds"]])

        main_dept = safe_str(row[WARD_COLS["main_dept"]])
        new_patients_yr = safe_int(row[WARD_COLS["new_patients_yr"]])
        build_year = safe_str(row[WARD_COLS["build_year"]])

        total_permitted_beds = general_beds + therapy_beds

        ward_data = {
            "ward_name": ward_name,
            "ward_code": safe_str(row[WARD_COLS["ward_code"]]),
            "function_r5": function_r5,
            "function_r7": function_r7,
            "general_beds_permitted": general_beds,
            "general_beds_max_use": general_max,
            "general_beds_min_use": general_min,
            "therapy_beds_permitted": therapy_beds,
            "therapy_beds_max_use": therapy_max,
            "therapy_beds_min_use": therapy_min,
            "therapy_medical_beds": therapy_med_beds,
            "total_permitted_beds": total_permitted_beds,
            "admission_fee": admission_fee,
            "admission_fee_beds": admission_fee_beds,
            "special_fee_1": special_fee_1,
            "special_fee_1_beds": special_fee_1_beds,
            "special_fee_2": special_fee_2,
            "special_fee_2_beds": special_fee_2_beds,
            "nursing_ratio_estimated": estimate_nursing_ratio(
                admission_fee, general_beds, therapy_beds, nurse_ft, nurse_pt
            ),
            "nurse_ft": nurse_ft,
            "nurse_pt": nurse_pt,
            "prac_nurse_ft": prac_nurse_ft,
            "prac_nurse_pt": prac_nurse_pt,
            "aide_ft": aide_ft,
            "aide_pt": aide_pt,
            "pt_ft": pt_ft,
            "pt_pt": pt_pt,
            "ot_ft": ot_ft,
            "ot_pt": ot_pt,
            "st_ft": st_ft,
            "st_pt": st_pt,
            "main_dept": main_dept,
            "new_patients_yearly": new_patients_yr,
            "build_year": build_year,
        }

        h["wards"].append(ward_data)

    wb.close()

    print(f"  Total rows scanned: {row_count}")
    print(f"  Kanagawa rows: {kanagawa_count}")
    print(f"  Target area rows (wards): {target_count}")
    print(f"  Unique hospitals: {len(hospitals)}")

    return dict(hospitals)


def parse_facility_file():
    """Parse the facility-level Excel file for additional hospital info."""
    print(f"\nLoading facility file: {FACILITY_FILE}")

    wb = openpyxl.load_workbook(FACILITY_FILE, read_only=True)
    ws = wb.active

    facilities = {}

    for row in ws.iter_rows(min_row=7, max_col=25, values_only=True):
        pref_code = safe_str(row[FACILITY_COLS["pref_code"]])
        if pref_code != KANAGAWA_PREF_CODE:
            continue

        city_name = safe_str(row[FACILITY_COLS["city_name"]])
        if city_name not in NISHI_KANAGAWA_CITIES:
            continue

        medical_code = safe_str(row[FACILITY_COLS["medical_code"]])
        facilities[medical_code] = {
            "type": safe_str(row[FACILITY_COLS["type"]]),
            "owner": safe_str(row[FACILITY_COLS["owner"]]),
            "dpc_group": safe_str(row[FACILITY_COLS["dpc_group"]]),
            "tokutei": safe_str(row[FACILITY_COLS["tokutei"]]),
            "chiiki_shien": safe_str(row[FACILITY_COLS["chiiki_shien"]]),
            "sogo_nyuin": safe_str(row[FACILITY_COLS["sogo_nyuin"]]),
            "kyuseiki_juujitsu": safe_str(row[FACILITY_COLS["kyuseiki_juujitsu"]]),
            "zaitaku_shien": safe_str(row[FACILITY_COLS["zaitaku_shien"]]),
            "zaitaku_kouhou": safe_str(row[FACILITY_COLS["zaitaku_kouhou"]]),
        }

    wb.close()
    print(f"  Target facilities found: {len(facilities)}")
    return facilities


def build_output(hospitals, facilities):
    """Build the structured output JSON."""
    output = {}

    for code, h in sorted(hospitals.items(), key=lambda x: x[1]["city_name"]):
        # Create a short name (remove corporate prefix for display)
        short_name = h["name"]
        for prefix in ["医療法人社団", "医療法人財団", "医療法人", "社会医療法人",
                        "社会福祉法人", "独立行政法人", "公益社団法人", "公益財団法人",
                        "社会ｼﾞｬﾊﾟﾝﾒﾃﾞｨｶﾙｱﾗｲｱﾝｽ", "国立研究開発法人",
                        "学校法人", "宗教法人"]:
            if prefix in short_name:
                # Keep the part after the corporate name
                pass  # Keep full name; short_name is just for the key

        # Aggregate ward-level data
        total_nurse_ft = sum(w["nurse_ft"] for w in h["wards"])
        total_nurse_pt = sum(w["nurse_pt"] for w in h["wards"])
        total_permitted = sum(w["total_permitted_beds"] for w in h["wards"])
        total_max_use = sum(w["general_beds_max_use"] + w["therapy_beds_max_use"]
                           for w in h["wards"])
        functions = sorted(set(w["function_r5"] for w in h["wards"] if w["function_r5"]))
        admission_fees = sorted(set(w["admission_fee"] for w in h["wards"]
                                    if w["admission_fee"] and w["admission_fee"] != "-"))

        # Get facility info
        fac = facilities.get(code, {})

        hospital_entry = {
            "medical_code": code,
            "name": h["name"],
            "city_name": h["city_name"],
            "city_code": h["city_code"],
            "secondary_area": h["secondary_area_name"],
            "vision_area": h["vision_area_name"],
            "ward_count": len(h["wards"]),
            "total_permitted_beds": total_permitted,
            "total_max_use_beds": total_max_use,
            "total_nurse_ft": total_nurse_ft,
            "total_nurse_pt": total_nurse_pt,
            "functions": functions,
            "admission_fees": admission_fees,
            "facility_type": fac.get("type", ""),
            "owner": fac.get("owner", ""),
            "dpc_group": fac.get("dpc_group", ""),
            "is_tokutei": fac.get("tokutei", "") == "あり",
            "is_chiiki_shien": fac.get("chiiki_shien", "") == "あり",
            "is_zaitaku_shien": fac.get("zaitaku_shien", "") == "あり",
            "wards": [],
        }

        for w in h["wards"]:
            ward_entry = {
                "ward_name": w["ward_name"],
                "function_current": w["function_r5"],
                "function_planned_2025": w["function_r7"],
                "admission_fee": w["admission_fee"],
                "admission_fee_beds": w["admission_fee_beds"],
                "nursing_ratio": w["nursing_ratio_estimated"],
                "general_beds_permitted": w["general_beds_permitted"],
                "general_beds_max_use": w["general_beds_max_use"],
                "therapy_beds_permitted": w["therapy_beds_permitted"],
                "therapy_beds_max_use": w["therapy_beds_max_use"],
                "therapy_medical_beds": w["therapy_medical_beds"],
                "total_permitted_beds": w["total_permitted_beds"],
                "nurse_ft": w["nurse_ft"],
                "nurse_pt": w["nurse_pt"],
                "prac_nurse_ft": w["prac_nurse_ft"],
                "prac_nurse_pt": w["prac_nurse_pt"],
                "aide_ft": w["aide_ft"],
                "aide_pt": w["aide_pt"],
                "rehab_staff": {
                    "pt_ft": w["pt_ft"],
                    "pt_pt": w["pt_pt"],
                    "ot_ft": w["ot_ft"],
                    "ot_pt": w["ot_pt"],
                    "st_ft": w["st_ft"],
                    "st_pt": w["st_pt"],
                },
                "main_dept": w["main_dept"],
                "new_patients_yearly": w["new_patients_yearly"],
                "build_year": w["build_year"],
            }

            # Add special fees if present
            if w["special_fee_1"]:
                ward_entry["special_fee_1"] = w["special_fee_1"]
                ward_entry["special_fee_1_beds"] = w["special_fee_1_beds"]
            if w["special_fee_2"]:
                ward_entry["special_fee_2"] = w["special_fee_2"]
                ward_entry["special_fee_2_beds"] = w["special_fee_2_beds"]

            hospital_entry["wards"].append(ward_entry)

        output[code] = hospital_entry

    return output


def compare_with_existing(parsed_data, existing_json_path):
    """Compare parsed data with existing kanagawa_ward_data.json."""
    print(f"\nComparing with existing: {existing_json_path}")

    if not os.path.exists(existing_json_path):
        print("  WARNING: Existing JSON file not found. Skipping comparison.")
        return None

    with open(existing_json_path, "r", encoding="utf-8") as f:
        existing = json.load(f)

    # Build lookup by medical code from existing data
    existing_by_code = {}
    for key, val in existing.items():
        code = val.get("medicalCode", "")
        if code:
            existing_by_code[code] = {"key": key, "data": val}

    # Build lookup by medical code from parsed data
    parsed_by_code = {}
    for code, val in parsed_data.items():
        parsed_by_code[code] = val

    comparison = {
        "summary": {
            "parsed_hospitals": len(parsed_data),
            "existing_hospitals": len(existing),
            "matched_by_code": 0,
            "only_in_parsed": 0,
            "only_in_existing": 0,
        },
        "discrepancies": [],
        "only_in_parsed": [],
        "only_in_existing": [],
        "matched_hospitals": [],
    }

    # Check each parsed hospital against existing
    for code, parsed_h in parsed_by_code.items():
        if code in existing_by_code:
            comparison["summary"]["matched_by_code"] += 1
            ex = existing_by_code[code]["data"]
            ex_key = existing_by_code[code]["key"]

            diffs = []

            # Compare ward count
            ex_ward_count = ex.get("wardCount", 0)
            parsed_ward_count = parsed_h["ward_count"]
            if ex_ward_count != parsed_ward_count:
                diffs.append({
                    "field": "ward_count",
                    "existing": ex_ward_count,
                    "parsed": parsed_ward_count,
                })

            # Compare functions
            ex_functions = sorted(ex.get("functions", []))
            parsed_functions = sorted(parsed_h["functions"])
            if ex_functions != parsed_functions:
                diffs.append({
                    "field": "functions",
                    "existing": ex_functions,
                    "parsed": parsed_functions,
                })

            # Compare total nurse count
            ex_total_nurses = ex.get("totalWardNurses", 0)
            parsed_total_nurses = parsed_h["total_nurse_ft"] + parsed_h["total_nurse_pt"]
            if abs(ex_total_nurses - parsed_total_nurses) > 1:
                diffs.append({
                    "field": "total_nurses",
                    "existing": ex_total_nurses,
                    "parsed": parsed_total_nurses,
                    "parsed_detail": f"FT:{parsed_h['total_nurse_ft']} + PT:{parsed_h['total_nurse_pt']}",
                })

            # Compare ward-level beds
            for i, pw in enumerate(parsed_h["wards"]):
                if i < len(ex.get("wards", [])):
                    ew = ex["wards"][i]
                    ex_beds = ew.get("beds", 0)
                    # parsed beds: use admission_fee_beds as the comparable metric
                    # or total_permitted_beds
                    parsed_beds = pw["total_permitted_beds"]
                    parsed_fee_beds = pw["admission_fee_beds"]

                    if ex_beds != parsed_beds and ex_beds != parsed_fee_beds:
                        diffs.append({
                            "field": f"ward[{i}].beds ({pw['ward_name']})",
                            "existing": ex_beds,
                            "parsed_permitted": parsed_beds,
                            "parsed_fee_beds": parsed_fee_beds,
                        })

            if diffs:
                comparison["discrepancies"].append({
                    "hospital": parsed_h["name"],
                    "existing_key": ex_key,
                    "medical_code": code,
                    "city": parsed_h["city_name"],
                    "differences": diffs,
                })
            else:
                comparison["matched_hospitals"].append({
                    "hospital": parsed_h["name"],
                    "medical_code": code,
                    "city": parsed_h["city_name"],
                    "status": "MATCH",
                })
        else:
            comparison["summary"]["only_in_parsed"] += 1
            comparison["only_in_parsed"].append({
                "hospital": parsed_h["name"],
                "medical_code": code,
                "city": parsed_h["city_name"],
                "ward_count": parsed_h["ward_count"],
                "total_beds": parsed_h["total_permitted_beds"],
                "functions": parsed_h["functions"],
            })

    # Check for hospitals in existing but not in parsed
    for code, ex_info in existing_by_code.items():
        if code not in parsed_by_code:
            comparison["summary"]["only_in_existing"] += 1
            comparison["only_in_existing"].append({
                "hospital": ex_info["data"].get("name", ex_info["key"]),
                "medical_code": code,
                "city": ex_info["data"].get("cityName", ""),
                "existing_key": ex_info["key"],
            })

    return comparison


def print_kobayashi_report(parsed_data):
    """Print detailed report for 小林病院."""
    print("\n" + "=" * 70)
    print("  DETAILED REPORT: 小林病院 (Kobayashi Hospital)")
    print("=" * 70)

    found = False
    for code, h in parsed_data.items():
        if "小林病院" in h["name"]:
            found = True
            print(f"\n  Medical Code: {h['medical_code']}")
            print(f"  Full Name: {h['name']}")
            print(f"  City: {h['city_name']} (code: {h['city_code']})")
            print(f"  Secondary Area: {h['secondary_area']}")
            print(f"  Vision Area: {h['vision_area']}")
            print(f"  Owner: {h.get('owner', 'N/A')}")
            print(f"  Ward Count: {h['ward_count']}")
            print(f"  Total Permitted Beds: {h['total_permitted_beds']}")
            print(f"  Total Max Use Beds: {h['total_max_use_beds']}")
            print(f"  Functions: {', '.join(h['functions'])}")
            print(f"  Total Nurses FT: {h['total_nurse_ft']}")
            print(f"  Total Nurses PT: {h['total_nurse_pt']}")
            print(f"  Admission Fees: {', '.join(h['admission_fees'])}")

            for i, w in enumerate(h["wards"]):
                print(f"\n  --- Ward {i+1}: {w['ward_name']} ---")
                print(f"    Function (R5 current): {w['function_current']}")
                print(f"    Function (R7 planned): {w['function_planned_2025']}")
                print(f"    Admission Fee: {w['admission_fee']}")
                print(f"    Admission Fee Beds: {w['admission_fee_beds']}")
                print(f"    Nursing Ratio (estimated): {w['nursing_ratio']}")
                print(f"    General Beds: permitted={w['general_beds_permitted']}, "
                      f"max_use={w['general_beds_max_use']}, "
                      f"therapy={w['therapy_beds_permitted']}")
                print(f"    Therapy Medical Beds: {w['therapy_medical_beds']}")
                print(f"    Nurses FT: {w['nurse_ft']}, PT: {w['nurse_pt']}")
                print(f"    Prac Nurses FT: {w['prac_nurse_ft']}, PT: {w['prac_nurse_pt']}")
                print(f"    Aides FT: {w['aide_ft']}, PT: {w['aide_pt']}")
                rehab = w.get("rehab_staff", {})
                print(f"    Rehab: PT={rehab.get('pt_ft',0)}+{rehab.get('pt_pt',0)}, "
                      f"OT={rehab.get('ot_ft',0)}+{rehab.get('ot_pt',0)}, "
                      f"ST={rehab.get('st_ft',0)}+{rehab.get('st_pt',0)}")
                print(f"    Main Dept: {w['main_dept']}")
                print(f"    New Patients/Year: {w['new_patients_yearly']}")
                print(f"    Build Year: {w['build_year']}")

    if not found:
        print("  小林病院 NOT FOUND in parsed data!")


def print_comparison_report(comparison):
    """Print a summary of the comparison."""
    if comparison is None:
        return

    s = comparison["summary"]
    print("\n" + "=" * 70)
    print("  COMPARISON REPORT: Parsed vs Existing kanagawa_ward_data.json")
    print("=" * 70)
    print(f"\n  Parsed hospitals: {s['parsed_hospitals']}")
    print(f"  Existing hospitals: {s['existing_hospitals']}")
    print(f"  Matched by code: {s['matched_by_code']}")
    print(f"  Only in parsed (NEW): {s['only_in_parsed']}")
    print(f"  Only in existing (MISSING): {s['only_in_existing']}")

    if comparison["discrepancies"]:
        print(f"\n  DISCREPANCIES FOUND: {len(comparison['discrepancies'])}")
        print("  " + "-" * 60)
        for d in comparison["discrepancies"]:
            print(f"\n  {d['hospital']} ({d['city']}) [code: {d['medical_code']}]")
            for diff in d["differences"]:
                field = diff["field"]
                if "existing" in diff and "parsed" in diff:
                    print(f"    {field}:")
                    print(f"      existing: {diff['existing']}")
                    print(f"      parsed:   {diff['parsed']}")
                    if "parsed_detail" in diff:
                        print(f"      detail:   {diff['parsed_detail']}")
                elif "parsed_permitted" in diff:
                    print(f"    {field}:")
                    print(f"      existing:         {diff['existing']}")
                    print(f"      parsed_permitted: {diff['parsed_permitted']}")
                    print(f"      parsed_fee_beds:  {diff['parsed_fee_beds']}")

    if comparison["only_in_parsed"]:
        print(f"\n  NEW HOSPITALS (in parsed but not in existing):")
        print("  " + "-" * 60)
        for h in comparison["only_in_parsed"]:
            print(f"    {h['hospital']} ({h['city']}) - {h['ward_count']} wards, "
                  f"{h['total_beds']} beds, functions: {', '.join(h['functions'])}")

    if comparison["only_in_existing"]:
        print(f"\n  MISSING (in existing but not in parsed):")
        print("  " + "-" * 60)
        for h in comparison["only_in_existing"]:
            print(f"    {h['hospital']} ({h['city']}) [code: {h['medical_code']}]")

    # Print full match list
    if comparison["matched_hospitals"]:
        matched = [h for h in comparison["matched_hospitals"]]
        if matched:
            print(f"\n  FULLY MATCHED: {len(matched)} hospitals (no discrepancies)")


def print_summary_by_city(parsed_data):
    """Print summary grouped by city."""
    print("\n" + "=" * 70)
    print("  SUMMARY BY CITY")
    print("=" * 70)

    by_city = defaultdict(list)
    for code, h in parsed_data.items():
        by_city[h["city_name"]].append(h)

    for city in sorted(by_city.keys()):
        hospitals = by_city[city]
        total_beds = sum(h["total_permitted_beds"] for h in hospitals)
        total_nurses = sum(h["total_nurse_ft"] + h["total_nurse_pt"] for h in hospitals)

        # Count functions
        func_count = defaultdict(int)
        for h in hospitals:
            for f in h["functions"]:
                func_count[f] += 1

        print(f"\n  {city}: {len(hospitals)} hospitals, {total_beds} total beds, "
              f"{total_nurses:.0f} total nurses")
        for h in hospitals:
            beds_str = f"{h['total_permitted_beds']}beds"
            funcs = "/".join(h["functions"])
            print(f"    - {h['name']} ({beds_str}, {funcs})")


def main():
    print("=" * 70)
    print("  病床機能報告 令和5年度 Parser")
    print("  Target: 神奈川県西部 hospitals")
    print("=" * 70)

    # Step 1: Parse ward file
    hospitals = parse_ward_file()

    # Step 2: Parse facility file
    facilities = parse_facility_file()

    # Step 3: Build structured output
    output = build_output(hospitals, facilities)

    # Step 4: Print 小林病院 report
    print_kobayashi_report(output)

    # Step 5: Compare with existing
    comparison = compare_with_existing(output, EXISTING_JSON)

    # Step 6: Print comparison report
    print_comparison_report(comparison)

    # Step 7: Print city summary
    print_summary_by_city(output)

    # Step 8: Save output
    print(f"\nSaving parsed data to: {OUTPUT_JSON}")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    if comparison:
        print(f"Saving comparison to: {COMPARISON_JSON}")
        with open(COMPARISON_JSON, "w", encoding="utf-8") as f:
            json.dump(comparison, f, ensure_ascii=False, indent=2)

    print(f"\nDone. {len(output)} hospitals parsed.")
    return output, comparison


if __name__ == "__main__":
    main()
